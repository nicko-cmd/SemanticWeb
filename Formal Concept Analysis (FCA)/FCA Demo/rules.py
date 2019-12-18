from openpyxl import load_workbook
from apyori import apriori
import re


def generate_rules(filename):
    """Generate association rules from formal concept.

    To extract the association rules from the formal concept
    it is used the 'apriori' algorithm.

    Arguments:
        filename (str): The file with formal concept
    """
    wb = load_workbook(filename)
    ws = wb.worksheets[0]

    trans = create_transactions(ws)

    # Obtain associations rules as records
    results = list(apriori(trans))
    rules = format(results)
    save(rules, filename)

    print('Rules generated.')


def create_transactions(ws):
    """Create the transactions from formal concept.

    A transaction is made by a list of only
    attributes where the binary relation is equal to 1.

    Arguments:
        ws (Worksheet): The formal concept

    Returns:
        transactions (list): A list of the transactions
    """
    transactions = []
    dim = ws.calculate_dimension()

    # Take last value cell
    a = re.split(':', dim)[1]

    # Take last attribute cell
    b = re.findall('[A-Z]+', a)[0] + '1'

    attributes = ws['B1':b]
    cells = ws['B2':a]

    for i in range(len(cells)):
        tmp = []
        for j in range(len(cells[i])):
            # Check if binary relation is equal to 1
            if int(cells[i][j].value) == 1:
                tmp.append(attributes[0][j].value)
        transactions.append(tmp)
    return transactions


def format(results):
    """Formatt associantions rules in easy way to read.

    Arguments:
        results (list): A list of Reletional Record

    Returns:
        rules (list): A list of formatted rules

    ReletionalRecord: tuple(items: set,
                            support: float,
                            ordered_statitics: list of
                            OrderedStatistic objects)

    OrderedStatistic: tuple(items_base: set,
                            items_add: set,
                            confidence: float,
                            lift: float)
    """
    items, os = 0, 2
    items_base, items_add, conf = 0, 1, 2
    rules = []

    for RR in results:
        # jump rules with one item
        if len(RR[items]) > 1:
            ord_sts = RR[os]
            for OS in ord_sts:
                confidence = round(OS[conf],2)
                itemsA = list(OS[items_add])
                itemsB = list(OS[items_base])

                # rule format: [conf=x] itemB => itemA
                rule = '(conf={}) {} => {}'.format(confidence, itemsB, itemsA)
                rules.append(rule)

    return rules


def save(rules, filename):
    """Save the rules in a text file.

    Arguments:
        rules (list): The list of the association rules
    """
    path = '/data/rules'
    file_rules = path + filename.replace('.xlsx', '.txt')

    with open(file_rules, 'w') as file:
        for rule in rules:
            file.write(rule + '\n')
